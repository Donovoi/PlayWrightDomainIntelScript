from openpyxl import Workbook, load_workbook
from filelock import FileLock
from openpyxl import load_workbook, Workbook
from lxml import etree
from threading import Thread
import queue
import os
import logging
import time
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError, Error as PlaywrightError
import cudf  # Import cuDF for GPU DataFrame processinga
from openpyxl.drawing.image import Image
from bs4 import BeautifulSoup  # Use BeautifulSoup to parse HTML content
from concurrent.futures import ThreadPoolExecutor, as_completed

logging.basicConfig(level=logging.DEBUG)


def create_dir(directory: str) -> str:
    """
    Create a directory if it doesn't exist.

    Args:
        directory (str): The directory path.

    Returns:
        str: The directory path.
    """
    if not os.path.exists(directory):
        os.makedirs(directory)
    return directory


def sanitize_filename(filename: str) -> str:
    """
    Sanitize a filename by removing or replacing invalid characters.

    Args:
        filename (str): The original filename.

    Returns:
        str: The sanitized filename.
    """
    return "".join(c for c in filename if c.isalnum() or c in (' ', '.', '_')).rstrip()


def save_page_content(page, directory: str, url: str, tool_name: str, domain: str):
    """
    Save the screenshot of a page to an Excel file.

    Args:
        page: The page object from Playwright.
        directory (str): The directory where the content will be saved.
        url (str): The page URL.
        tool_name (str): The tool name.
        domain (str): The domain name.
    """
    try:
        if page.is_closed():
            logging.error(
                "Page is already closed for URL: %s. Cannot save content.", url)
            return

        page.wait_for_selector("body", timeout=60000)

        # Capture the screenshot
        screenshot_path = os.path.join(directory, f"{tool_name}.png")
        page.screenshot(path=screenshot_path, full_page=True)

        # Ensure the directory exists
        full_directory = create_dir(directory)

        # Create or load the Excel workbook
        file_path = os.path.join(full_directory, f"{domain}_{tool_name}.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove the default sheet

        if tool_name in wb.sheetnames:
            ws = wb[tool_name]
        else:
            ws = wb.create_sheet(tool_name)

        # Insert the screenshot into the Excel sheet
        img = Image(screenshot_path)
        ws.add_image(img, 'A1')

        wb.save(file_path)

        logging.info("Saved screenshot from tool '%s' for domain '%s' in directory '%s'.",
                     tool_name, domain, full_directory)
    except (OSError, ValueError, RuntimeError, PlaywrightTimeoutError, PlaywrightError) as e:
        logging.error("Failed to save content for URL '%s': %s", url, str(e))


def parse_xml_safely(xmlstring):
    """
    Parse XML string with lxml, recovering from errors if necessary.

    Args:
        xmlstring (str): The XML content as a string.

    Returns:
        lxml.etree.Element: The parsed XML element tree.
    """
    parser = etree.XMLParser(recover=True)
    try:
        return etree.fromstring(xmlstring, parser=parser)
    except etree.XMLSyntaxError as e:
        logging.error("Failed to parse XML: %s", e)
        return None


def find_png_files(directory):
    """
    Recursively find all .png files in the given directory.

    Args:
        directory (str): The directory to search.

    Returns:
        list: A list of file paths to .png files.
    """
    png_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.png'):
                png_files.append(os.path.join(root, file))
    return png_files


def save_to_excel(domain: str, tool_name: str, directory: str, content_type: str):
    """
    Save processed content to an Excel file.

    Args:
        domain (str): The domain name.
        tool_name (str): The tool name.
        directory (str): The directory to search for .png files.
        content_type (str): The type of content ('file_path' or 'html_content').
    """
    file_path = os.path.join("results", f"{domain}.xlsx")
    lock_path = f"{file_path}.lock"
    lock = FileLock(lock_path)

    try:
        with lock:
            if os.path.exists(file_path):
                try:
                    wb = load_workbook(file_path)
                except Exception as e:
                    logging.error("Failed to load workbook: %s", e)
                    # Attempt to recover from a malformed XML file by directly parsing it
                    with open(file_path, 'rb') as f:
                        xml_content = f.read()
                        recovered_tree = parse_xml_safely(xml_content)
                        if recovered_tree is None:
                            return  # Skip this file if recovery fails
                        # Potentially reconstruct workbook from recovered_tree here if needed
                        return
            else:
                wb = Workbook()
                wb.remove(wb.active)  # Remove the default sheet

            if tool_name in wb.sheetnames:
                ws = wb[tool_name]
            else:
                ws = wb.create_sheet(tool_name)

            if content_type == 'file_path':
                png_files = find_png_files(directory)
                row = 1
                for png_file in png_files:
                    if os.path.exists(png_file):
                        img = Image(png_file)
                        cell_position = f'A{row}'
                        ws.add_image(img, cell_position)
                        logging.info(
                            f"Added image {png_file} to {cell_position}")
                        row += 10  # Adjust row increment as needed
                    else:
                        logging.warning("PNG file not found: %s", png_file)
            elif content_type == 'html_content':
                for line in content:
                    ws.append([line])
                    logging.info(f"Added line to sheet: {line}")
            else:
                logging.error("Invalid content type: %s", content_type)

            wb.save(file_path)
            logging.info(f"Workbook saved to {file_path}")
    except Exception as e:
        logging.error("An error occurred while saving to Excel: %s", e)


def parse_html_file(file_path: str) -> str:
    """
    Parse the HTML content from a file.

    Args:
        file_path (str): The path to the HTML file.

    Returns:
        str: The text content of the HTML file.
    """
    with open(file_path, "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, 'html.parser')
        text_content = soup.get_text(separator='\n')
    return text_content


def process_html_file(file_path: str, domain: str, tool_name: str):
    """
    Process an HTML file and save the processed content to an Excel file.

    Args:
        file_path (str): The path to the HTML file.
        domain (str): The domain name.
        tool_name (str): The tool name.
    """
    try:
        content = parse_html_file(file_path)

        # Process content using cuDF
        content_df = cudf.Series(content.splitlines())
        filtered_content = content_df.str.strip().dropna()

        # Save processed content to Excel
        save_to_excel(domain, tool_name,
                      filtered_content.to_arrow().to_pylist(), content_type='file_path')

        logging.info(
            "Saved content from tool '%s' for domain '%s'.", tool_name, domain)
    except (OSError, ValueError, RuntimeError) as e:
        logging.error("Failed to save content for file '%s': %s",
                      file_path, str(e))


def worker(file_queue):
    while True:
        try:
            file_path, domain, tool_name = file_queue.get(timeout=1)
            process_html_file(file_path, domain, tool_name)
        except queue.Empty:
            break


def recurse_and_process(base_directory: str):
    file_queue = queue.Queue()

    for root, dirs, files in os.walk(base_directory):
        domain = os.path.basename(root)
        if not dirs:  # Leaf directory
            for file in files:
                if file.endswith(".html"):
                    tool_name = file.split('_')[0]
                    file_path = os.path.join(root, file)
                    file_queue.put((file_path, domain, tool_name))

    threads = []
    for _ in range(40):  # Max 40 threads
        thread = Thread(target=worker, args=(file_queue,))
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()


def trigger_all_searches(page, domain: str, search_functions: list):
    """
    Trigger all search functions on a page.

    Args:
        page: The page object from Playwright.
        domain (str): The domain name.
        search_functions (list): A list of search function names to trigger.
    """
    logging.info("Triggering all search functions for domain: %s", domain)

    for search_function in search_functions:
        function_exists = page.evaluate(
            f"typeof {search_function} === 'function'")
        if not function_exists:
            logging.warning(
                "Function %s is not defined on the page.", search_function)
            continue

        page.evaluate(f"""
        (function() {{
            try {{
                var domain = '{domain}';
                {search_function}(domain);
            }} catch (e) {{
                console.error("Error: " + e.message);
            }}
        }})();
        """)

        time.sleep(0.5)


def capture_pages(context, domain: str, max_retries: int = 3):
    """
    Capture content from all pages in a context with retries.

    Args:
        context: The browser context object from Playwright.
        domain (str): The domain name.
        max_retries (int): The maximum number of retries for capturing content.
    """
    pages = context.pages[1:]  # Exclude the first page (main page)
    for page in pages:
        retries = max_retries
        tool_name = page.title()  # Use the page title as the tool name
        while retries > 0:
            try:
                if page.is_closed():
                    logging.warning(
                        "Page is closed unexpectedly. Skipping further retries.")
                    break

                page.wait_for_load_state()
                logging.info("Capturing content for page: %s", page.url)

                if domain in page.url or domain in page.content():
                    save_page_content(page, create_dir(
                        domain), page.url, tool_name, domain)
                    break
                else:
                    logging.warning(
                        "Search term '%s' not found on the page or URL is incorrect: %s", domain, page.url)
                    retries -= 1
                    if retries > 0:
                        logging.info("Retrying... (%d retries left)", retries)
                        if not page.is_closed():
                            page.close()
                        new_page = context.new_page()
                        new_page.goto(page.url, timeout=60000)
                        page = new_page
            except PlaywrightError as e:
                if "ERR_CONNECTION_REFUSED" in str(e):
                    logging.error(
                        "Connection refused for page: %s. Skipping...", page.url)
                    break
                else:
                    logging.warning(
                        "Timeout or error while loading page: %s", page.url)
                    retries -= 1
                    if retries > 0 and not page.is_closed():
                        logging.info("Retrying... (%d retries left)", retries)
                        page.reload()
            finally:
                if not page.is_closed():
                    page.close()

        if retries == 0:
            logging.error(
                "Failed to capture content after retries for page: %s", page.url)


def process_domain(domain: str, playwright):
    """
    Process a domain by navigating to the Intel Techniques page and capturing content.

    Args:
        domain (str): The domain name.
        playwright: The Playwright object for browser automation.
    """
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()

    try:
        page = context.new_page()
        logging.info(
            "Navigating to the Intel Techniques page for domain: %s", domain)

        page.goto("https://inteltechniques.com/tools/Domain.html", timeout=60000)
        page.wait_for_load_state('domcontentloaded')

        if page.url == "about:blank":
            logging.error("Page failed to load, URL is still 'about:blank'")
            return

        logging.info("Successfully loaded the page: %s", page.url)

        search_functions = [f"doSearch{i:02d}" for i in range(1, 69)]

        trigger_all_searches(page, domain, search_functions)

        page.wait_for_timeout(20000)

        capture_pages(context, domain)

    except PlaywrightTimeoutError as e:
        logging.error("Failed to load the page: %s", str(e))
    finally:
        context.close()
        browser.close()


def run_playwright(pw):
    """
    Run Playwright to process domains from a file using ThreadPoolExecutor for concurrency.

    Args:
        pw: The Playwright object for browser automation.
    """
    with open("urls.txt", "r", encoding="utf-8") as f:
        domains = f.read().splitlines()

    with ThreadPoolExecutor(max_workers=40) as executor:
        futures = [executor.submit(process_domain, domain, pw)
                   for domain in domains]

        for future in as_completed(futures):
            try:
                future.result()  # Check for exceptions
            except Exception as exc:
                logging.error(
                    "An error occurred while processing a domain: %s", exc)


def main():
    """
    Main function to execute the script.
    """
    folder_path = input(
        "Enter the folder path containing local HTML files (press Enter to use URLs): ")

    if folder_path and os.path.isdir(folder_path):
        logging.info(
            "Processing local HTML files in directory: %s", folder_path)
        recurse_and_process(folder_path)
    else:
        logging.info(
            "No valid folder path provided. Using URLs from urls.txt.")
        with sync_playwright() as playwright:
            run_playwright(playwright)


if __name__ == "__main__":
    main()
